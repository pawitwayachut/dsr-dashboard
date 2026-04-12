#!/usr/bin/env python3
"""
DSR Dashboard Generator
Reads Excel source files → generates self-contained HTML dashboard
Usage: python generate_dashboard.py
"""

import openpyxl
import json
import sys
import struct
import io
import calendar as cal_mod
import zipfile
from pathlib import Path
from datetime import datetime, date, timedelta

# ─── PATHS ────────────────────────────────────────────────────────────────────

SCRIPT_DIR  = Path(__file__).parent
BASE        = SCRIPT_DIR / "OPR - Daily Sales Report"
now         = datetime.now()
month_label = now.strftime("%b%Y")
OUTPUT_FILE = SCRIPT_DIR / f"DSR_Dashboard_{month_label}.html"

THAI_DOW   = ['จ.', 'อ.', 'พ.', 'พฤ.', 'ศ.', 'ส.', 'อา.']
THAI_HOLIDAYS = {
    (2026,4,6): 'Chakri', (2026,4,13): 'Songkran',
    (2026,4,14): 'Songkran', (2026,4,15): 'Songkran',
    (2026,5,1): 'Labour Day', (2026,5,4): 'Coronation Day',
    (2026,6,3): 'Queen Birthday',
}

# ─── FILE DISCOVERY ───────────────────────────────────────────────────────────

def find_latest_folder(base_path):
    folders = sorted([
        d for d in Path(base_path).iterdir()
        if d.is_dir() and d.name.isdigit() and len(d.name) == 6
    ])
    if not folders:
        raise FileNotFoundError(f"No YYYYMM folders found in {base_path}")
    return folders[-1]

ssp_folder  = find_latest_folder(BASE / "SSP")
mono_folder = find_latest_folder(BASE / "MONO")

# Safe fallback directory (outside OneDrive, immune to sync corruption)
SAFE_DIR = Path("/sessions/clever-vigilant-cannon")

def repair_zip_file(filepath):
    """Repair xlsx files truncated by OneDrive (missing EOCD). Returns repaired path or None."""
    with open(filepath, 'rb') as f:
        data = f.read()
    # Check if already valid
    try:
        zipfile.ZipFile(io.BytesIO(data)).close()
        return None  # already valid
    except Exception:
        pass
    entries = []
    pos = 0
    while True:
        idx = data.find(b'PK\x03\x04', pos)
        if idx == -1: break
        entries.append(idx)
        pos = idx + 4
    if not entries: return None
    output = io.BytesIO()
    cd_entries = []
    for offset in entries:
        if offset + 30 > len(data): break
        sig, ver, flags, method, mtime, mdate, crc32, comp_size, uncomp_size, name_len, extra_len = \
            struct.unpack_from('<4sHHHHHIIIHH', data, offset)
        if sig != b'PK\x03\x04': break
        header_size = 30 + name_len + extra_len
        filename = data[offset+30:offset+30+name_len]
        extra = data[offset+30+name_len:offset+30+name_len+extra_len]
        if comp_size == 0 and (flags & 0x08):
            next_pk = data.find(b'PK', offset + header_size + 1)
            if next_pk == -1: next_pk = len(data)
            file_data = data[offset+header_size:next_pk]
            dd_pos = file_data.rfind(b'PK\x07\x08')
            if dd_pos != -1: file_data = file_data[:dd_pos]
            comp_size = len(file_data)
        else:
            file_data = data[offset+header_size:offset+header_size+comp_size]
        local_offset = output.tell()
        output.write(data[offset:offset+30+name_len+extra_len])
        output.write(file_data)
        cd_entries.append({
            'ver': ver, 'flags': flags & ~0x08, 'method': method,
            'mtime': mtime, 'mdate': mdate, 'crc32': crc32,
            'comp_size': comp_size, 'uncomp_size': uncomp_size if uncomp_size else comp_size,
            'name': filename, 'extra': extra, 'offset': local_offset
        })
    cd_offset = output.tell()
    for e in cd_entries:
        output.write(struct.pack('<4sHHHHHHIIIHHHHHII',
            b'PK\x01\x02', 20, e['ver'], e['flags'], e['method'],
            e['mtime'], e['mdate'], e['crc32'], e['comp_size'], e['uncomp_size'],
            len(e['name']), len(e['extra']), 0, 0, 0, 0x20, e['offset']))
        output.write(e['name'])
        output.write(e['extra'])
    cd_size = output.tell() - cd_offset
    output.write(struct.pack('<4sHHHHIIH', b'PK\x05\x06', 0, 0,
        len(cd_entries), len(cd_entries), cd_size, cd_offset, 0))
    output.seek(0)
    try:
        zipfile.ZipFile(output).close()
    except Exception:
        return None
    repaired_path = SCRIPT_DIR / f".repaired_{filepath.stem}.xlsx"
    output.seek(0)
    with open(repaired_path, 'wb') as f:
        f.write(output.read())
    print(f"  Repaired truncated file: {filepath.name} → {repaired_path}")
    return repaired_path

def find_readable(primary_glob, safe_pattern=None):
    """Find first readable xlsx file from primary glob, fallback to repair or safe dir."""
    for p in primary_glob:
        try:
            zipfile.ZipFile(p).close()
            return p
        except Exception:
            # Try repairing
            repaired = repair_zip_file(p)
            if repaired:
                return repaired
    if safe_pattern:
        for p in sorted(SAFE_DIR.glob(safe_pattern)):
            try:
                zipfile.ZipFile(p).close()
                print(f"  Using safe fallback: {p.name}")
                return p
            except Exception:
                pass
    return next(iter(primary_glob), None)

bkk_path    = find_readable(list(ssp_folder.glob("ROM_BKK_SALE*.xlsx")), "ROM_BKK_SAFE*.xlsx")
upc_path    = find_readable(list(ssp_folder.glob("ROM_UPC_SALE*.xlsx")), "ROM_UPC_SAFE*.xlsx")
mono_path_raw = mono_folder / "Daily Analyst.xlsx"
try:
    zipfile.ZipFile(mono_path_raw).close()
    mono_path = mono_path_raw
except Exception:
    repaired = repair_zip_file(mono_path_raw)
    mono_path = repaired if repaired else mono_path_raw

# SSP MTD Sales Tracking file — lives in SSP/YYYYMM folder
# Provides: daily targets (Summary sheet) + LY daily sales (Daily Sales sheet)
tracking_candidates = list(ssp_folder.glob("SSP MTD Sales Tracking*.xlsx"))
if tracking_candidates:
    trk_raw = tracking_candidates[0]
    try:
        zipfile.ZipFile(trk_raw).close()
        daily_summary_path = trk_raw
    except Exception:
        repaired = repair_zip_file(trk_raw)
        daily_summary_path = repaired if repaired else trk_raw
else:
    daily_summary_path = None

print("Files found:")
print(f"  BKK : {bkk_path.name}")
print(f"  UPC : {upc_path.name}")
print(f"  MONO: {mono_path.name}")
if daily_summary_path:
    print(f"  TRK : {daily_summary_path.name}")
print()

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def safe(val, default=0):
    if val is None: return default
    if isinstance(val, str) and (val.startswith('#') or val == ''): return default
    return val

def fmt_m(n):
    if n is None: return '—'
    n = float(n)
    if abs(n) >= 1_000_000: return f"{n/1e6:.1f}M"
    if abs(n) >= 1_000: return f"{n/1e3:.0f}K"
    return f"{n:.0f}"

def fmt_m2(n):
    """Millions with 2 decimals, no suffix."""
    if n is None or n == 0: return ''
    return f"{float(n)/1e6:.2f}"

def fmt_pct_html(val, good_positive=True):
    if val is None: return '<span class="neutral">—</span>'
    pct = val * 100
    cls = 'pos' if ((pct >= 0) == good_positive) else 'neg'
    sign = '+' if pct >= 0 else ''
    return f'<span class="{cls}">{sign}{pct:.1f}%</span>'

def fmt_diff_cell(diff_val):
    """Format diff in M with background color. Zero = no highlight."""
    if diff_val is None: return '<td class="num"></td>'
    v = diff_val / 1e6
    if abs(v) < 0.005:
        return f'<td class="num">{v:.2f}</td>'
    cls = 'cell-pos' if v > 0 else 'cell-neg'
    return f'<td class="num {cls}">{v:.2f}</td>'

def fmt_chg_cell(actual, ly):
    """Format %Chg cell with background color. Zero = no highlight."""
    if actual is None or ly is None or ly == 0: return '<td class="num"></td>'
    pct = (actual / ly - 1) * 100
    sign = '+' if pct >= 0 else ''
    if abs(pct) < 0.05:
        return f'<td class="num">{sign}{pct:.1f}%</td>'
    cls = 'cell-pos' if pct > 0 else 'cell-neg'
    return f'<td class="num {cls}">{sign}{pct:.1f}%</td>'

def thai_date(dt):
    return f"{THAI_DOW[dt.weekday()]}, {dt.strftime('%d/%m/%Y')}"

def total(lst, key):
    return sum(s.get(key) or 0 for s in lst)

# ─── DAILY SUMMARY PARSER (SSP MTD Sales Tracking) ─────────────────────────

def parse_daily_summary(filepath):
    """Parse the Summary sheet for per-day targets, actuals, and forecast landing.
    Returns: (daily_targets dict {day_num: target},
              daily_actuals dict {day_num: actual},
              forecast_landing, budget)
    """
    if filepath is None:
        return {}, {}, None, None
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"  WARNING: Cannot read tracking file ({e}), skipping")
        return {}, {}, None, None
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
    else:
        ws = wb[wb.sheetnames[0]]
    daily_tg = {}
    daily_act = {}
    forecast_landing = None
    budget = None
    for row in ws.iter_rows(min_row=3, max_col=8, values_only=True):
        col_a = row[0]
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None

        if isinstance(col_c, datetime) and isinstance(col_d, (int, float)):
            day_num = col_c.day
            daily_tg[day_num] = col_d
            if isinstance(col_e, (int, float)) and col_e > 0:
                daily_act[day_num] = col_e

        if col_a == 'Y25 Sales' or (isinstance(col_d, str) and 'Forecast' in col_d):
            if isinstance(col_e, (int, float)):
                forecast_landing = col_e

        if isinstance(col_d, str) and 'Budget' in col_d:
            if isinstance(col_e, (int, float)):
                budget = col_e

    wb.close()
    return daily_tg, daily_act, forecast_landing, budget

# ─── SSP PARSER ───────────────────────────────────────────────────────────────

def parse_ssp(filepath, region):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    numbered = sorted(
        [int(s) for s in wb.sheetnames if s.isdigit() and 1 <= int(s) <= 31],
        reverse=True
    )
    active_ws = None
    data_day  = None
    for day in numbered:
        ws = wb[str(day)]
        for row in ws.iter_rows(min_row=4, max_row=30, values_only=True):
            code = row[0]
            if code and isinstance(code, str) and 1 <= len(code) <= 4 and code not in ('Store Code',):
                daily = safe(row[5], None)
                if daily is not None and isinstance(daily, (int, float)) and daily > 0:
                    active_ws, data_day = ws, day
                    break
        if active_ws: break

    if not active_ws:
        print(f"  WARNING: No data found in {filepath.name}")
        return [], [], None

    print(f"  {region}: Latest data = day {data_day}")

    # Two-pass: first collect all rows, then assign DM names correctly
    # In Excel, stores come BEFORE their DM total row
    raw_rows = []
    for row in active_ws.iter_rows(min_row=4, values_only=True):
        code = row[0]
        if not code: continue
        if code in ('Store Code', 'Target day'): continue
        if isinstance(code, str) and (code.startswith('Total') or code.startswith('Target')): continue
        is_dm = isinstance(code, str) and 'Total' in code
        dm_name = str(code).replace('Total','').strip() if is_dm else None
        raw_rows.append((row, is_dm, dm_name))

    # Pass 2: assign DM by looking FORWARD to the next DM total row
    stores, dm_totals = [], []
    for i, (row, is_dm, dm_name) in enumerate(raw_rows):
        if is_dm:
            # This is the DM total — stores above belong to this DM
            continue
        # Find the next DM total row after this store
        assigned_dm = "—"
        for j in range(i + 1, len(raw_rows)):
            if raw_rows[j][1]:  # is_dm
                assigned_dm = raw_rows[j][2]
                break

        entry = {
            'code'     : str(row[0]),
            'name'     : str(row[1]) if row[1] else str(row[0]),
            'dm'       : assigned_dm,
            'region'   : region,
            'daily_ty' : safe(row[5]),
            'daily_ly' : safe(row[8]),
            'daily_pct': safe(row[9], None),
            'mtd_ty'   : safe(row[10]),
            'mtd_ly'   : safe(row[11]),
            'mtd_pct'  : safe(row[12], None),
            'target'   : safe(row[13]),
            'pct_ach'  : safe(row[15], None),
            'avg_wd'   : safe(row[2]),
            'avg_fri'  : safe(row[3]),
            'avg_wknd' : safe(row[4]),
        }
        stores.append(entry)

    # Build DM totals from the DM rows
    for row, is_dm, dm_name in raw_rows:
        if not is_dm: continue
        entry = {
            'code'     : dm_name,
            'name'     : dm_name,
            'dm'       : dm_name,
            'region'   : region,
            'daily_ty' : safe(row[5]),
            'daily_ly' : safe(row[8]),
            'daily_pct': safe(row[9], None),
            'mtd_ty'   : safe(row[10]),
            'mtd_ly'   : safe(row[11]),
            'mtd_pct'  : safe(row[12], None),
            'target'   : safe(row[13]),
            'pct_ach'  : safe(row[15], None),
            'avg_wd'   : safe(row[2]),
            'avg_fri'  : safe(row[3]),
            'avg_wknd' : safe(row[4]),
        }
        dm_totals.append(entry)

    return stores, dm_totals, data_day

# ─── LY DAILY LOOKUP (from SSP MTD Sales Tracking → Daily Sales sheet) ─────

def parse_ly_daily_from_tracking(filepath):
    """Parse 'Daily Sales' sheet for LY daily totals.
    Returns dict: {date -> total_sales}
    """
    if filepath is None:
        return {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return {}
    if 'Daily Sales' not in wb.sheetnames:
        return {}
    ws = wb['Daily Sales']
    daily = {}
    for row in ws.iter_rows(min_row=6, max_col=4, values_only=True):
        dt = row[0]
        if isinstance(dt, datetime):
            dt = dt.date()
        elif not isinstance(dt, date):
            continue
        if dt.year != now.year - 1:
            continue
        sales = row[1]
        if isinstance(sales, (int, float)) and sales > 0:
            daily[dt] = sales
    return daily

# ─── DAILY HISTORY PARSER ────────────────────────────────────────────────────

def parse_daily_history(filepath):
    """Parse ALL day sheets (1-31) for daily TY actual + LY totals."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    ly_date_map = {}

    for day in range(1, 32):
        if str(day) not in wb.sheetnames: continue
        ws = wb[str(day)]

        # Get LY date from header row 1, col 8
        hdr = list(ws.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True))[0]
        if isinstance(hdr[8], datetime):
            ly_date_map[day] = hdr[8].date()

        day_ty, day_ly = 0.0, 0.0
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = row[0]
            if not code or not isinstance(code, str): continue
            if code in ('Store Code', 'Target day'): continue
            if 'Total' in code or code.startswith('Target'): continue
            ty = row[5] if isinstance(row[5], (int, float)) else 0
            ly = row[8] if isinstance(row[8], (int, float)) else 0
            day_ty += ty
            day_ly += ly

        result[day] = {
            'actual': day_ty if day_ty > 0 else None,
            'ly': day_ly if day_ly > 0 else None,
        }
    return result, ly_date_map

# ─── MONO PARSER ──────────────────────────────────────────────────────────────

MONTH_ABBR = {1:'JAN',2:'FEB',3:'MAR',4:'APR',5:'MAY',6:'JUN',
              7:'JUL',8:'AUG',9:'SEP',10:'OCT',11:'NOV',12:'DEC'}

def find_mono_sheet(wb):
    m, y = MONTH_ABBR[now.month], str(now.year)
    for name in wb.sheetnames:
        u = name.strip().upper()
        if 'DAILY ANALYST' in u and m in u and y in u: return wb[name]
    for name in reversed(wb.sheetnames):
        if 'DAILY ANALYST' in name.upper(): return wb[name]
    return None

def parse_mono(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = find_mono_sheet(wb)
    if not ws:
        print("  WARNING: Cannot find current month sheet in MONO file")
        return [], None
    print(f"  MONO: Using sheet '{ws.title.strip()}'")
    date_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    day_cols = [(i, v.date()) for i, v in enumerate(date_row) if isinstance(v, datetime)]
    if not day_cols:
        print("  WARNING: No date columns found in MONO sheet")
        return [], None
    all_rows = list(ws.iter_rows(min_row=4, values_only=True))
    latest_col = latest_date = None
    for col, dt in reversed(day_cols):
        for row in all_rows:
            if row[0] is not None:
                ty = row[col + 1]
                if ty is not None and ty != 0:
                    latest_col, latest_date = col, dt
                    break
        if latest_col is not None: break
    print(f"  MONO: Latest data date = {latest_date}")
    stores = []
    for row in all_rows:
        if row[0] is None: continue
        brand, store_name = row[2], row[3]
        if not brand or not store_name: continue
        if latest_col is not None:
            d_ly, d_ty = safe(row[latest_col]), safe(row[latest_col+1])
            d_pct = safe(row[latest_col+2], None)
        else:
            d_ly = d_ty = 0; d_pct = None
        mtd_ty = mtd_ly = 0.0
        for col, dt in day_cols:
            if latest_col is not None and col > latest_col: break
            if row[col+1]: mtd_ty += row[col+1]
            if row[col]:   mtd_ly += row[col]
        mtd_pct = (mtd_ty/mtd_ly - 1) if mtd_ly else None
        stores.append({
            'brand': str(brand), 'name': str(store_name),
            'daily_ty': d_ty, 'daily_ly': d_ly, 'daily_pct': d_pct,
            'mtd_ty': round(mtd_ty,2), 'mtd_ly': round(mtd_ly,2), 'mtd_pct': mtd_pct,
        })
    return stores, latest_date

# ═══════════════════════════════════════════════════════════════════════════════
# PARSE ALL DATA
# ═══════════════════════════════════════════════════════════════════════════════

print("Parsing BKK stores...")
bkk_stores, bkk_dms, bkk_day = parse_ssp(bkk_path, 'BKK')

print("Parsing UPC stores...")
upc_stores, upc_dms, upc_day = parse_ssp(upc_path, 'UPC')

print("Parsing daily history BKK...")
bkk_hist, bkk_ly_dates = parse_daily_history(bkk_path)
print("Parsing daily history UPC...")
upc_hist, upc_ly_dates = parse_daily_history(upc_path)

print("Parsing LY daily (from SSP MTD Sales Tracking)...")
ly_daily_from_tracking = parse_ly_daily_from_tracking(daily_summary_path)
if ly_daily_from_tracking:
    print(f"  LY daily: {len(ly_daily_from_tracking)} days from tracking file")

print("Parsing MONO...")
mono_stores, mono_date = parse_mono(mono_path)

# ─── COMBINED DAILY ───────────────────────────────────────────────────────────

data_day_num = bkk_day or now.day
month_days   = cal_mod.monthrange(now.year, now.month)[1]

combined_daily = {}
for d in range(1, month_days + 1):
    bh = bkk_hist.get(d, {})
    uh = upc_hist.get(d, {})
    ba, ua = bh.get('actual') or 0, uh.get('actual') or 0
    bl, ul = bh.get('ly') or 0, uh.get('ly') or 0
    combined_daily[d] = {
        'bkk_ty': bh.get('actual'), 'upc_ty': uh.get('actual'),
        'actual': (ba + ua) if (ba or ua) else None,
        'ly': (bl + ul) if (bl or ul) else None,
    }

# LY date for each TY day (use BKK dates, confirmed DOW-aligned)
ly_dates = bkk_ly_dates

# ─── TOTALS & METRICS ─────────────────────────────────────────────────────────

bkk_mtd    = total(bkk_stores, 'mtd_ty')
upc_mtd    = total(upc_stores, 'mtd_ty')
ssp_mtd    = bkk_mtd + upc_mtd
mono_mtd   = total(mono_stores, 'mtd_ty')
group_mtd  = ssp_mtd + mono_mtd

bkk_mtd_ly   = total(bkk_stores, 'mtd_ly')
upc_mtd_ly   = total(upc_stores, 'mtd_ly')
ssp_mtd_ly   = bkk_mtd_ly + upc_mtd_ly
mono_mtd_ly  = total(mono_stores, 'mtd_ly')
group_mtd_ly = ssp_mtd_ly + mono_mtd_ly

# ─── Read daily targets from summary file ────────────────────────────────────
print("Parsing daily summary targets...")
daily_targets, daily_actuals_from_summary, summary_forecast, summary_budget = parse_daily_summary(daily_summary_path)
if daily_targets:
    print(f"  Found {len(daily_targets)} daily targets from {daily_summary_path.name}")
    # Cache to JSON for reliability
    cache_path = ssp_folder / "daily_targets_cache.json"
    with open(cache_path, 'w') as cf:
        json.dump({
            "month": now.strftime("%Y-%m"),
            "daily_targets": {str(k): v for k, v in daily_targets.items()},
            "forecast_landing": summary_forecast,
            "budget": summary_budget
        }, cf, indent=2)
    print(f"  Cached targets to {cache_path.name}")
else:
    # Try JSON cache fallback
    cache_path = ssp_folder / "daily_targets_cache.json"
    if cache_path.exists():
        with open(cache_path) as cf:
            cache = json.load(cf)
        if cache.get("month") == now.strftime("%Y-%m"):
            daily_targets = {int(k): v for k, v in cache["daily_targets"].items()}
            summary_forecast = cache.get("forecast_landing")
            summary_budget = cache.get("budget")
            print(f"  Loaded {len(daily_targets)} daily targets from cache (JSON)")
        else:
            print("  WARNING: Cache is for a different month, falling back to store-level targets")
    else:
        print("  WARNING: No daily summary file or cache found, falling back to store-level targets")

# Monthly target: prefer summary file forecast, fallback to store-level sum
bkk_target = total(bkk_stores, 'target')
upc_target = total(upc_stores, 'target')
store_level_target = bkk_target + upc_target
ssp_monthly_target = summary_forecast if summary_forecast else store_level_target

ssp_pct_ly  = (ssp_mtd / ssp_mtd_ly - 1) if ssp_mtd_ly else None
mono_pct_ly = (mono_mtd / mono_mtd_ly - 1) if mono_mtd_ly else None
group_pct_ly = (group_mtd / group_mtd_ly - 1) if group_mtd_ly else None

# Daily totals
ssp_daily  = total(bkk_stores, 'daily_ty') + total(upc_stores, 'daily_ty')
mono_daily = total(mono_stores, 'daily_ty')
group_daily = ssp_daily + mono_daily

# If no daily summary targets, fall back to weighted calculation from store averages
if not daily_targets:
    ssp_avg_wd   = sum(safe(s.get('avg_wd'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_wd'),0),(int,float)))
    ssp_avg_fri  = sum(safe(s.get('avg_fri'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_fri'),0),(int,float)))
    ssp_avg_wknd = sum(safe(s.get('avg_wknd'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_wknd'),0),(int,float)))
    weighted_total = 0
    for d in range(1, month_days + 1):
        dow = date(now.year, now.month, d).weekday()
        if dow == 4: weighted_total += ssp_avg_fri
        elif dow >= 5: weighted_total += ssp_avg_wknd
        else: weighted_total += ssp_avg_wd
    scale = ssp_monthly_target / weighted_total if weighted_total else 0
    for d in range(1, month_days + 1):
        dow = date(now.year, now.month, d).weekday()
        if dow == 4: daily_targets[d] = ssp_avg_fri * scale
        elif dow >= 5: daily_targets[d] = ssp_avg_wknd * scale
        else: daily_targets[d] = ssp_avg_wd * scale

# Today's daily target
today_target = daily_targets.get(data_day_num, 0)

# MTD target (sum of daily targets 1..data_day)
mtd_target = sum(daily_targets.get(d,0) for d in range(1, data_day_num + 1))
mtd_vs_tg  = (ssp_mtd / mtd_target - 1) if mtd_target else None

# LY full month — sum all LY daily from tracking file
ly_full_month_est = sum(ly_daily_from_tracking.values())

# LY MTD — sum DOW-aligned LY days from tracking file
ty_d1_dow = date(now.year, now.month, 1).weekday()
ly_d1_dow = date(now.year-1, now.month, 1).weekday()
_offset = ty_d1_dow - ly_d1_dow
if _offset < 0: _offset += 7
ly_mtd = 0
for d in range(1, data_day_num + 1):
    ly_day = d + _offset
    ly_month_days_total = cal_mod.monthrange(now.year-1, now.month)[1]
    if ly_day <= ly_month_days_total:
        ly_dt = date(now.year-1, now.month, ly_day)
        ly_mtd += ly_daily_from_tracking.get(ly_dt, 0)

# Forecast Landing = monthly target
forecast_landing = ssp_monthly_target
forecast_vs_ly   = (forecast_landing / ly_full_month_est - 1) if ly_full_month_est else None

# ─── RUNRATE CALCULATION ─────────────────────────────────────────────────────
# Use store-level avg weekday/weekend to project remaining days
ssp_avg_wd_all   = sum(safe(s.get('avg_wd'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_wd'),0),(int,float)))
ssp_avg_wknd_all = sum(safe(s.get('avg_wknd'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_wknd'),0),(int,float)))
ssp_avg_fri_all  = sum(safe(s.get('avg_fri'),0) for s in bkk_stores + upc_stores if isinstance(safe(s.get('avg_fri'),0),(int,float)))

remaining_wd = 0
remaining_wknd = 0
for d in range(data_day_num + 1, month_days + 1):
    dow = date(now.year, now.month, d).weekday()
    if dow >= 5:
        remaining_wknd += 1
    elif dow == 4:
        remaining_wd += 1  # Count Friday as weekday (uses avg_fri but close enough, or separate)
    else:
        remaining_wd += 1

# Use avg weekday for Mon-Fri, avg weekend for Sat-Sun
remaining_projection = 0
for d in range(data_day_num + 1, month_days + 1):
    dow = date(now.year, now.month, d).weekday()
    if dow >= 5:
        remaining_projection += ssp_avg_wknd_all
    elif dow == 4:
        remaining_projection += ssp_avg_fri_all
    else:
        remaining_projection += ssp_avg_wd_all

runrate = ssp_mtd + remaining_projection
runrate_vs_tg = (runrate / ssp_monthly_target - 1) if ssp_monthly_target else None
runrate_vs_ly = (runrate / ly_full_month_est - 1) if ly_full_month_est else None
runrate_diff_tg = runrate - ssp_monthly_target

# ─── DAILY KPI METRICS ──────────────────────────────────────────────────────
# Daily: LY daily for same TY day
ssp_daily_ly = total(bkk_stores, 'daily_ly') + total(upc_stores, 'daily_ly')
daily_diff = ssp_daily - today_target
daily_vs_tg = (ssp_daily / today_target - 1) if today_target else None
daily_vs_ly = (ssp_daily / ssp_daily_ly - 1) if ssp_daily_ly else None

# MTD diffs
mtd_diff = ssp_mtd - mtd_target
mtd_vs_tg = (ssp_mtd / mtd_target - 1) if mtd_target else None

# BKK metrics — vs TG uses MTD target (prorated from SSP MTD target by region weight)
bkk_full_tg = total(bkk_stores, 'target')
upc_full_tg = total(upc_stores, 'target')
total_full_tg = bkk_full_tg + upc_full_tg
bkk_mtd_tg = mtd_target * (bkk_full_tg / total_full_tg) if total_full_tg else 0
upc_mtd_tg = mtd_target * (upc_full_tg / total_full_tg) if total_full_tg else 0
bkk_diff = bkk_mtd - bkk_mtd_ly
bkk_vs_tg = (bkk_mtd / bkk_mtd_tg - 1) if bkk_mtd_tg else None
bkk_vs_ly = (bkk_mtd / bkk_mtd_ly - 1) if bkk_mtd_ly else None

# UPC metrics
upc_diff = upc_mtd - upc_mtd_ly
upc_vs_tg = (upc_mtd / upc_mtd_tg - 1) if upc_mtd_tg else None
upc_vs_ly = (upc_mtd / upc_mtd_ly - 1) if upc_mtd_ly else None

# MONO metrics — use LY as target proxy (no separate target file)
mono_diff = mono_mtd - mono_mtd_ly
mono_vs_tg = (mono_mtd / mono_mtd_ly - 1) if mono_mtd_ly else None

# Progress bar
progress_pct = (ssp_mtd / ssp_monthly_target * 100) if ssp_monthly_target else 0
mtd_tg_pct = (mtd_target / ssp_monthly_target * 100) if ssp_monthly_target else 0

# Labels
data_day_label = f"{now.strftime('%b')} {data_day_num}, {now.year}" if bkk_day else now.strftime("%b %d, %Y")
generated_at   = now.strftime("%d %b %Y %H:%M")

print(f"\nTotals:")
print(f"  SSP  MTD : {ssp_mtd:>15,.0f}   vs LY {(ssp_pct_ly or 0)*100:+.1f}%")
print(f"  MONO MTD : {mono_mtd:>15,.0f}   vs LY {(mono_pct_ly or 0)*100:+.1f}%")
print(f"  Monthly TG: {ssp_monthly_target:>14,.0f}")
print(f"  Runrate:    {runrate:>14,.0f}")
print(f"  LY Full Mo: {ly_full_month_est:>14,.0f}\n")

# ═══════════════════════════════════════════════════════════════════════════════
# HTML BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Daily Tracking Table ─────────────────────────────────────────────────────

def build_daily_tracking():
    yr, mo = now.year, now.month
    ty_d1_dow = date(yr, mo, 1).weekday()
    ly_d1_dow = date(yr-1, mo, 1).weekday()
    offset = ty_d1_dow - ly_d1_dow
    if offset < 0: offset += 7

    ly_month_days = cal_mod.monthrange(yr-1, mo)[1]
    rows = []

    # Orphan LY rows (no TY match) — get actual LY sales from tracking file
    for ld in range(1, offset + 1):
        ly_dt = date(yr-1, mo, ld)
        ly_sale = ly_daily_from_tracking.get(ly_dt, 0)
        ly_str = fmt_m2(ly_sale) if ly_sale > 0 else '—'
        rows.append(
            f'<tr class="orphan-row">'
            f'<td>{thai_date(ly_dt)}</td>'
            f'<td class="num">{ly_str}</td>'
            f'<td></td><td></td><td></td><td></td><td></td><td></td></tr>'
        )

    # Group into weeks (Mon-Sun by TY date)
    week_buf = []
    week_num = 0
    current_iso_week = None

    def flush_week(buf, wn):
        if not buf: return ''
        w_ly  = sum(r[1] for r in buf if r[1])
        w_tg  = sum(r[2] for r in buf if r[2])
        w_act = sum(r[3] for r in buf if r[3])
        has_actual = any(r[3] for r in buf)
        w_diff = (w_act - w_tg) if has_actual else None
        ly_str = fmt_m2(w_ly) if w_ly else ''
        tg_str = fmt_m2(w_tg) if w_tg else ''
        act_str = fmt_m2(w_act) if has_actual else ''
        if w_diff is not None:
            diff_m = w_diff / 1e6
            if abs(round(diff_m, 2)) == 0:
                diff_cell = f'<td class="num">0.00</td>'
            else:
                diff_cls = 'cell-pos' if diff_m > 0 else 'cell-neg'
                diff_cell = f'<td class="num {diff_cls}">{diff_m:.2f}</td>'
        else:
            diff_cell = '<td class="num"></td>'
        if has_actual and w_ly > 0:
            w_chg = (w_act / w_ly - 1) * 100
            if abs(round(w_chg, 1)) == 0:
                chg_cell = f'<td class="num">0.0%</td>'
            else:
                chg_cls = 'cell-pos' if w_chg > 0 else 'cell-neg'
                chg_sign = '+' if w_chg > 0 else ''
                chg_cell = f'<td class="num {chg_cls}">{chg_sign}{w_chg:.1f}%</td>'
        else:
            chg_cell = '<td class="num"></td>'
        return (
            f'<tr class="week-row">'
            f'<td>{thai_date(date(yr-1, mo, min(buf[0][4], ly_month_days)))}</td>'
            f'<td class="num"><strong>{ly_str}</strong></td>'
            f'<td><strong>W{wn}</strong></td>'
            f'<td class="num"><strong>{tg_str}</strong></td>'
            f'<td class="num"><strong>{act_str}</strong></td>'
            f'{diff_cell}'
            f'{chg_cell}'
            f'<td></td></tr>'
        )

    for d in range(1, month_days + 1):
        ty_dt = date(yr, mo, d)
        iso_wk = ty_dt.isocalendar()[1]

        if current_iso_week is not None and iso_wk != current_iso_week:
            week_num += 1
            rows.append(flush_week(week_buf, week_num))
            week_buf = []
        current_iso_week = iso_wk

        cd = combined_daily.get(d, {})
        tg_val = daily_targets.get(d, 0)
        summary_act = daily_actuals_from_summary.get(d)
        # If reforecast (target == Summary actual), use Summary actual so diff = 0
        if summary_act and tg_val and abs(tg_val - summary_act) < 1:
            actual = summary_act
        else:
            actual = cd.get('actual')

        # LY date (DOW-aligned) — LY from tracking file
        ly_day = d + offset
        if ly_day <= ly_month_days:
            ly_dt = date(yr-1, mo, ly_day)
            ly_date_str = thai_date(ly_dt)
            ly_val = ly_daily_from_tracking.get(ly_dt, 0) or None
        else:
            ly_date_str = ''
            ly_val = None

        # Holiday remark
        hol = THAI_HOLIDAYS.get((yr, mo, d), '')
        remark_html = f'<span class="tag-holiday">{hol}</span>' if hol else ''

        # Is this after latest data? (future)
        is_future = d > data_day_num
        row_cls = 'future-row' if is_future else ''
        if hol: row_cls += ' holiday-row'

        # Cells
        ly_sales_str = fmt_m2(ly_val) if ly_val else ''
        tg_str = fmt_m2(tg_val) if tg_val else ''

        if actual:
            act_str = fmt_m2(actual)
            diff_val = actual - tg_val
            diff_m = diff_val / 1e6
            if abs(round(diff_m, 2)) == 0:
                diff_str = f'<td class="num">0.00</td>'
            else:
                diff_cls = 'cell-pos' if diff_val > 0 else 'cell-neg'
                diff_str = f'<td class="num {diff_cls}">{diff_m:.2f}</td>'
            chg_pct = (actual / ly_val - 1) * 100 if (ly_val and ly_val > 0) else None
            if chg_pct is not None:
                if abs(round(chg_pct, 1)) == 0:
                    chg_str = f'<td class="num">0.0%</td>'
                else:
                    chg_cls = 'cell-pos' if chg_pct > 0 else 'cell-neg'
                    chg_sign = '+' if chg_pct > 0 else ''
                    chg_str = f'<td class="num {chg_cls}">{chg_sign}{chg_pct:.1f}%</td>'
        else:
            act_str = ''
            diff_str = '<td class="num"></td>'
            chg_str = '<td class="num"></td>'

        # Store for week subtotal: (ty_day, ly_val_raw, tg_raw, actual_raw, ly_day)
        week_buf.append((d, ly_val or 0, tg_val, actual or 0 if actual else 0, ly_day))

        rows.append(
            f'<tr class="{row_cls}">'
            f'<td>{ly_date_str}</td>'
            f'<td class="num">{ly_sales_str}</td>'
            f'<td>{thai_date(ty_dt)}</td>'
            f'<td class="num">{tg_str}</td>'
            f'<td class="num">{act_str}</td>'
            f'{diff_str}'
            f'{chg_str}'
            f'<td>{remark_html}</td></tr>'
        )

    # Flush last week
    week_num += 1
    rows.append(flush_week(week_buf, week_num))

    # ── Footer rows ──
    # MTD Sales
    mtd_ly_m = ly_mtd / 1e6
    mtd_ty_m = ssp_mtd / 1e6
    mtd_chg = (ssp_mtd / ly_mtd - 1) * 100 if ly_mtd else 0
    mtd_chg_cls = 'cell-pos' if mtd_chg >= 0 else 'cell-neg'
    rows.append(
        f'<tr class="summary-row">'
        f'<td><strong>MTD Sales</strong></td>'
        f'<td class="num"><strong>{mtd_ly_m:.2f}</strong></td>'
        f'<td></td>'
        f'<td><strong>Sales Landing</strong></td>'
        f'<td class="num"><strong>{mtd_ty_m:.2f}</strong></td>'
        f'<td></td>'
        f'<td class="num {mtd_chg_cls}">{mtd_chg:+.1f}%</td>'
        f'<td></td></tr>'
    )

    # Forecast Landing
    fc_val = ssp_monthly_target / 1e6
    ly_full = ly_full_month_est / 1e6
    fc_chg = (ssp_monthly_target / ly_full_month_est - 1) * 100 if ly_full_month_est else 0
    fc_chg_cls = 'cell-pos' if fc_chg >= 0 else 'cell-neg'
    rows.append(
        f'<tr class="summary-row forecast-row">'
        f'<td><strong>Y25 Sales</strong></td>'
        f'<td class="num"><strong>{ly_full:.2f}</strong></td>'
        f'<td></td>'
        f'<td><strong>Forecast Landing</strong></td>'
        f'<td class="num highlight-yellow"><strong>{fc_val:.2f}</strong></td>'
        f'<td></td>'
        f'<td class="num {fc_chg_cls}">{fc_chg:+.1f}%</td>'
        f'<td></td></tr>'
    )

    return '\n'.join(rows)

# ─── SSP Store Table (grouped by DM) ─────────────────────────────────────────

def build_ssp_table():
    from collections import OrderedDict

    def _region_subtotal(store_list, dm_list, region_label):
        """Build rows for one region (BKK or UPC) with DM groups + region subtotal."""
        dm_lookup = {d['dm']: d for d in dm_list}
        groups = OrderedDict()
        for s in store_list:
            dm = s['dm']
            if dm not in groups: groups[dm] = []
            groups[dm].append(s)

        region_cls = 'bkk' if region_label == 'BKK' else 'upc'
        badge = f'<span class="badge {region_cls}">{region_label}</span>'
        rows = []

        for dm, stores in groups.items():
            for s in stores:
                rows.append(
                    f'<tr data-region="{s["region"]}" data-dm="{dm}">'
                    f'<td class="code">{s["code"]}</td>'
                    f'<td class="store-name">{s["name"]}</td>'
                    f'<td>{badge}</td>'
                    f'<td class="num">{fmt_m(s["daily_ty"])}</td>'
                    f'<td class="num">{fmt_m(s["daily_ly"])}</td>'
                    f'<td class="num">{fmt_pct_html(s["daily_pct"])}</td>'
                    f'<td class="num">{fmt_m(s["mtd_ty"])}</td>'
                    f'<td class="num">{fmt_m(s["mtd_ly"])}</td>'
                    f'<td class="num">{fmt_pct_html(s["mtd_pct"])}</td>'
                    f'<td class="num">{fmt_m(s["target"])}</td>'
                    f'<td class="num">{fmt_pct_html(s["pct_ach"])}</td>'
                    f'</tr>'
                )
            # DM subtotal row
            d = dm_lookup.get(dm)
            if d:
                rows.append(
                    f'<tr class="dm-row" data-region="{region_label}" data-dm="{dm}">'
                    f'<td colspan="2"><strong>{d["name"]}</strong></td>'
                    f'<td>{badge}</td>'
                    f'<td class="num"><strong>{fmt_m(d["daily_ty"])}</strong></td>'
                    f'<td class="num">{fmt_m(d["daily_ly"])}</td>'
                    f'<td class="num">{fmt_pct_html(d["daily_pct"])}</td>'
                    f'<td class="num"><strong>{fmt_m(d["mtd_ty"])}</strong></td>'
                    f'<td class="num">{fmt_m(d["mtd_ly"])}</td>'
                    f'<td class="num">{fmt_pct_html(d["mtd_pct"])}</td>'
                    f'<td class="num">{fmt_m(d["target"])}</td>'
                    f'<td class="num">{fmt_pct_html(d["pct_ach"])}</td>'
                    f'</tr>'
                )

        # Region subtotal
        r_daily = total(store_list, 'daily_ty')
        r_daily_ly = total(store_list, 'daily_ly')
        r_mtd = total(store_list, 'mtd_ty')
        r_mtd_ly = total(store_list, 'mtd_ly')
        r_target = total(store_list, 'target')
        r_daily_pct = (r_daily / r_daily_ly - 1) if r_daily_ly else None
        r_mtd_pct = (r_mtd / r_mtd_ly - 1) if r_mtd_ly else None
        r_ach = (r_mtd / r_target) if r_target else None
        rows.append(
            f'<tr class="region-total-row" data-region="{region_label}">'
            f'<td colspan="2"><strong>Total {region_label}</strong></td>'
            f'<td>{badge}</td>'
            f'<td class="num"><strong>{fmt_m(r_daily)}</strong></td>'
            f'<td class="num"><strong>{fmt_m(r_daily_ly)}</strong></td>'
            f'<td class="num"><strong>{fmt_pct_html(r_daily_pct)}</strong></td>'
            f'<td class="num"><strong>{fmt_m(r_mtd)}</strong></td>'
            f'<td class="num"><strong>{fmt_m(r_mtd_ly)}</strong></td>'
            f'<td class="num"><strong>{fmt_pct_html(r_mtd_pct)}</strong></td>'
            f'<td class="num"><strong>{fmt_m(r_target)}</strong></td>'
            f'<td class="num"><strong>{fmt_pct_html(r_ach)}</strong></td>'
            f'</tr>'
        )
        return rows

    # Build BKK rows, then UPC rows
    all_rows = []
    all_rows.extend(_region_subtotal(bkk_stores, bkk_dms, 'BKK'))
    all_rows.extend(_region_subtotal(upc_stores, upc_dms, 'UPC'))

    # Grand Total (BKK + UPC)
    all_s = bkk_stores + upc_stores
    g_daily = total(all_s, 'daily_ty')
    g_daily_ly = total(all_s, 'daily_ly')
    g_mtd = total(all_s, 'mtd_ty')
    g_mtd_ly = total(all_s, 'mtd_ly')
    g_target = total(all_s, 'target')
    g_daily_pct = (g_daily / g_daily_ly - 1) if g_daily_ly else None
    g_mtd_pct = (g_mtd / g_mtd_ly - 1) if g_mtd_ly else None
    g_ach = (g_mtd / g_target) if g_target else None
    all_rows.append(
        f'<tr class="grand-total-row" data-region="ALL">'
        f'<td colspan="2"><strong>Grand Total SSP</strong></td>'
        f'<td></td>'
        f'<td class="num"><strong>{fmt_m(g_daily)}</strong></td>'
        f'<td class="num"><strong>{fmt_m(g_daily_ly)}</strong></td>'
        f'<td class="num"><strong>{fmt_pct_html(g_daily_pct)}</strong></td>'
        f'<td class="num"><strong>{fmt_m(g_mtd)}</strong></td>'
        f'<td class="num"><strong>{fmt_m(g_mtd_ly)}</strong></td>'
        f'<td class="num"><strong>{fmt_pct_html(g_mtd_pct)}</strong></td>'
        f'<td class="num"><strong>{fmt_m(g_target)}</strong></td>'
        f'<td class="num"><strong>{fmt_pct_html(g_ach)}</strong></td>'
        f'</tr>'
    )

    return '\n'.join(all_rows)

# ─── MONO Table (with brand subtotals) ───────────────────────────────────────

def build_mono_table():
    brand_totals = {}
    for s in mono_stores:
        b = s['brand']
        if b not in brand_totals:
            brand_totals[b] = {'brand':b, 'daily_ty':0,'daily_ly':0,'mtd_ty':0,'mtd_ly':0}
        for k in ('daily_ty','daily_ly','mtd_ty','mtd_ly'):
            brand_totals[b][k] += s.get(k) or 0

    for b in brand_totals:
        bt = brand_totals[b]
        bt['daily_pct'] = (bt['daily_ty']/bt['daily_ly']-1) if bt['daily_ly'] else None
        bt['mtd_pct']   = (bt['mtd_ty']/bt['mtd_ly']-1) if bt['mtd_ly'] else None

    rows = []
    prev_brand = None
    brand_buf = []

    def flush_brand(brand):
        bt = brand_totals.get(brand)
        if not bt: return
        rows.append(
            f'<tr class="brand-total" data-brand="{brand}">'
            f'<td colspan="2"><strong>{brand} Total</strong></td>'
            f'<td></td>'
            f'<td class="num"><strong>{fmt_m(bt["daily_ty"])}</strong></td>'
            f'<td class="num">{fmt_m(bt["daily_ly"])}</td>'
            f'<td class="num">{fmt_pct_html(bt["daily_pct"])}</td>'
            f'<td class="num"><strong>{fmt_m(bt["mtd_ty"])}</strong></td>'
            f'<td class="num">{fmt_m(bt["mtd_ly"])}</td>'
            f'<td class="num">{fmt_pct_html(bt["mtd_pct"])}</td>'
            f'</tr>'
        )

    for s in mono_stores:
        if s['brand'] != prev_brand:
            if prev_brand: flush_brand(prev_brand)
            prev_brand = s['brand']
        rows.append(
            f'<tr data-brand="{s["brand"]}">'
            f'<td class="code">—</td>'
            f'<td class="store-name">{s["name"]}</td>'
            f'<td><span class="badge mono-b">{s["brand"]}</span></td>'
            f'<td class="num">{fmt_m(s["daily_ty"])}</td>'
            f'<td class="num">{fmt_m(s["daily_ly"])}</td>'
            f'<td class="num">{fmt_pct_html(s["daily_pct"])}</td>'
            f'<td class="num">{fmt_m(s["mtd_ty"])}</td>'
            f'<td class="num">{fmt_m(s["mtd_ly"])}</td>'
            f'<td class="num">{fmt_pct_html(s["mtd_pct"])}</td>'
            f'</tr>'
        )
    if prev_brand: flush_brand(prev_brand)

    # Grand Total row
    gt = {'daily_ty':0,'daily_ly':0,'mtd_ty':0,'mtd_ly':0}
    for bt in brand_totals.values():
        for k in gt: gt[k] += bt[k]
    gt_daily_pct = (gt['daily_ty']/gt['daily_ly']-1) if gt['daily_ly'] else None
    gt_mtd_pct   = (gt['mtd_ty']/gt['mtd_ly']-1) if gt['mtd_ly'] else None
    rows.append(
        f'<tr class="grand-total-row">'
        f'<td colspan="3"><strong>MONO Grand Total</strong></td>'
        f'<td class="num"><strong>{fmt_m(gt["daily_ty"])}</strong></td>'
        f'<td class="num">{fmt_m(gt["daily_ly"])}</td>'
        f'<td class="num">{fmt_pct_html(gt_daily_pct)}</td>'
        f'<td class="num"><strong>{fmt_m(gt["mtd_ty"])}</strong></td>'
        f'<td class="num">{fmt_m(gt["mtd_ly"])}</td>'
        f'<td class="num">{fmt_pct_html(gt_mtd_pct)}</td>'
        f'</tr>'
    )
    return '\n'.join(rows)

# ─── Pre-computed HTML ────────────────────────────────────────────────────────

def ssp_filter_btns():
    return '\n'.join([
        '<button class="filter-btn active" onclick="filterSSP(\'ALL\',this)">All</button>',
        '<button class="filter-btn" onclick="filterSSP(\'BKK\',this)">BKK</button>',
        '<button class="filter-btn" onclick="filterSSP(\'UPC\',this)">UPC</button>',
    ])

def mono_filter_btns():
    btns = ['<button class="filter-btn active" onclick="filterMono(\'ALL\',this)">All Brands</button>']
    for b in sorted(set(s['brand'] for s in mono_stores)):
        btns.append(f'<button class="filter-btn" onclick="filterMono(\'{b}\',this)">{b}</button>')
    return '\n'.join(btns)

DAILY_TABLE_HTML  = build_daily_tracking()
SSP_TABLE_HTML    = build_ssp_table()
MONO_TABLE_HTML   = build_mono_table()
SSP_FILTER_BTNS   = ssp_filter_btns()
MONO_FILTER_BTNS  = mono_filter_btns()
mtd_diff_sign = '+' if mtd_diff >= 0 else ''

# ═══════════════════════════════════════════════════════════════════════════════
# ASSEMBLE HTML
# ═══════════════════════════════════════════════════════════════════════════════

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DSR Dashboard</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f5f5f5;color:#1a1a1a;font-size:13px}}
  .header{{background:#ffffff;border-bottom:3px solid #F90507;padding:12px 24px;display:flex;justify-content:space-between;align-items:center}}
  .header h1{{font-size:18px;font-weight:700;color:#002060}}
  .header .meta{{color:#5F615E;font-size:12px;text-align:right;line-height:1.7}}
  .header .meta strong{{color:#002060}}

  .kpi-bar{{display:flex;gap:1px;background:#e7e6e6}}
  .kpi-card{{flex:1;background:#ffffff;padding:12px 16px;border-bottom:3px solid transparent}}
  .kpi-card.daily{{border-color:#F90507}}
  .kpi-card.mtd{{border-color:#002060}}
  .kpi-card.forecast{{border-color:#183C92}}
  .kpi-card.bkk{{border-color:#F90507}}
  .kpi-card.upc{{border-color:#002060}}
  .kpi-card.mono{{border-color:#5F615E}}
  .kpi-label{{font-size:10px;color:#5F615E;text-transform:uppercase;letter-spacing:.6px;margin-bottom:2px}}
  .kpi-value{{font-size:20px;font-weight:700;color:#002060}}
  .kpi-sub{{font-size:11px;margin-top:2px;color:#5F615E}}

  .progress-section{{background:#ffffff;padding:8px 24px;display:flex;align-items:center;gap:14px;border-bottom:1px solid #e7e6e6}}
  .progress-label{{font-size:11px;color:#5F615E;white-space:nowrap;min-width:180px}}
  .progress-track{{flex:1;background:#e7e6e6;border-radius:4px;height:8px;position:relative}}
  .progress-fill{{height:100%;border-radius:4px;background:linear-gradient(90deg,#F90507,#183C92);transition:width .4s}}
  .mtd-tg-marker{{position:absolute;top:-4px;width:2px;height:16px;background:#F90507;border-radius:1px;transform:translateX(-1px)}}
  .progress-pct{{font-size:12px;font-weight:700;color:#002060;min-width:44px;text-align:right}}

  .tabs{{display:flex;background:#ffffff;border-bottom:1px solid #D9D9D9;padding:0 24px}}
  .tab-btn{{padding:11px 22px;cursor:pointer;font-size:13px;font-weight:500;color:#5F615E;border:none;background:none;border-bottom:2px solid transparent;transition:.15s}}
  .tab-btn:hover{{color:#002060}}
  .tab-btn.active{{color:#F90507;border-bottom-color:#F90507}}
  .tab-content{{display:none;padding:20px 24px}}
  .tab-content.active{{display:block}}
  .section-title{{font-size:13px;font-weight:600;color:#002060;margin:0 0 12px;padding-bottom:6px;border-bottom:2px solid #F90507;text-transform:uppercase;letter-spacing:.4px}}
  .grid-3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;margin-bottom:20px}}
  .card{{background:#ffffff;border-radius:8px;padding:16px;border:1px solid #e7e6e6}}
  .chart-wrap{{position:relative;height:200px}}

  .perf-region{{font-size:10px;color:#5F615E;min-width:30px}}
  .perf-val{{font-size:12px;font-weight:700;min-width:55px;text-align:right}}
  .perf-mtd{{font-size:11px;color:#5F615E;min-width:50px;text-align:right}}

  .filter-bar{{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}}
  .filter-btn{{padding:5px 14px;border-radius:20px;border:1px solid #D9D9D9;background:#ffffff;color:#5F615E;cursor:pointer;font-size:12px;transition:.15s}}
  .filter-btn:hover,.filter-btn.active{{background:#F90507;border-color:#F90507;color:#fff}}

  .table-wrap{{overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:12px}}
  th{{background:#002060;color:#ffffff;font-weight:600;text-transform:uppercase;font-size:10px;letter-spacing:.4px;padding:8px 10px;text-align:left;position:sticky;top:0;z-index:1}}
  td{{padding:6px 10px;border-bottom:1px solid #e7e6e6}}
  tr:hover td{{background:#f0f4ff}}
  td.num{{text-align:right;font-variant-numeric:tabular-nums}}
  td.code{{color:#5F615E;font-size:11px;font-family:monospace}}
  td.store-name{{font-weight:500;color:#1a1a1a}}
  .pos{{color:#0a8a3e}} .neg{{color:#d41920}} .neutral{{color:#5F615E}}

  /* DM subtotal rows — distinct style */
  tr.dm-row td{{background:#e8ecf4;font-weight:700;color:#002060;border-top:1px solid #b0c4de;border-bottom:2px solid #b0c4de;font-size:12px}}
  tr.dm-row:hover td{{background:#dce3f0}}
  tr.region-total-row td{{background:#ccd9ef;font-weight:700;color:#002060;border-top:2px solid #002060;border-bottom:3px solid #002060;font-size:13px}}
  tr.region-total-row:hover td{{background:#c0cfe8}}
  tr.grand-total-row td{{background:#fce8e8;font-weight:700;color:#F90507;border-top:2px solid #F90507;border-bottom:3px solid #F90507;font-size:13px}}
  tr.grand-total-row:hover td{{background:#f8d4d4}}

  /* Brand total rows */
  tr.brand-total td{{background:#e8ecf4;color:#002060;font-weight:700;border-top:1px solid #b0c4de;border-bottom:2px solid #b0c4de}}

  /* Daily tracking specific */
  tr.week-row td{{background:#e8ecf4;font-weight:700;color:#002060;border-top:2px solid #b0c4de;border-bottom:2px solid #b0c4de;font-size:12px}}
  tr.orphan-row td{{opacity:0.5;font-style:italic}}
  tr.future-row td{{opacity:0.4}}
  tr.holiday-row td{{background:#fff8e1}}
  tr.summary-row td{{background:#f0f4ff;font-weight:600;border-top:2px solid #D9D9D9}}
  tr.forecast-row td{{border-top:1px solid #D9D9D9}}
  .tag-holiday{{background:#F90507;color:#ffffff;font-size:10px;padding:1px 6px;border-radius:3px;margin-left:4px}}
  .cell-pos{{background:#e6f4ea;color:#0a8a3e}} .cell-neg{{background:#fce8e8;color:#d41920}}
  .highlight-yellow{{background:#fff3cd;color:#856404}}

  /* Daily Sales Tracking — all columns center-aligned */
  .daily-table th,.daily-table td{{text-align:center}}
  .daily-table td.num{{text-align:center}}

  .badge{{font-size:10px;padding:2px 7px;border-radius:3px;font-weight:600;letter-spacing:.3px}}
  .badge.bkk{{background:#fce8e8;color:#F90507}}
  .badge.upc{{background:#e0e8f5;color:#002060}}
  .badge.mono-b{{background:#e8ecf4;color:#183C92}}

  @media(max-width:900px){{
    .grid-2,.grid-3{{grid-template-columns:1fr}}
    .kpi-bar{{flex-wrap:wrap}}
    .kpi-card{{min-width:50%}}
  }}
</style>
</head>
<body>

<div class="header">
  <h1>OPR Daily Sales Report</h1>
  <div class="meta">
    <div>Data as of <strong>{data_day_label}</strong></div>
    <div>Generated {generated_at}</div>
  </div>
</div>

<!-- KPI BAR -->
<div class="kpi-bar">
  <div class="kpi-card daily">
    <div class="kpi-label">Daily (SSP)</div>
    <div class="kpi-value">{fmt_m(ssp_daily)}</div>
    <div class="kpi-sub">TG {fmt_m(today_target)} · Diff <span class="{'pos' if daily_diff>=0 else 'neg'}">{'+' if daily_diff>=0 else ''}{fmt_m(daily_diff)}</span> · vs TG {fmt_pct_html(daily_vs_tg)} · vs LY {fmt_pct_html(daily_vs_ly)}</div>
  </div>
  <div class="kpi-card mtd">
    <div class="kpi-label">MTD (SSP)</div>
    <div class="kpi-value">{fmt_m(ssp_mtd)}</div>
    <div class="kpi-sub">TG {fmt_m(mtd_target)} · Diff <span class="{'pos' if mtd_diff>=0 else 'neg'}">{'+' if mtd_diff>=0 else ''}{fmt_m(mtd_diff)}</span> · vs TG {fmt_pct_html(mtd_vs_tg)} · vs LY {fmt_pct_html(ssp_pct_ly)}</div>
  </div>
  <div class="kpi-card forecast">
    <div class="kpi-label">Runrate</div>
    <div class="kpi-value">{fmt_m(runrate)}</div>
    <div class="kpi-sub">LY {fmt_m(ly_full_month_est)} · Diff <span class="{'pos' if runrate_diff_tg>=0 else 'neg'}">{'+' if runrate_diff_tg>=0 else ''}{fmt_m(runrate_diff_tg)}</span> · vs TG {fmt_pct_html(runrate_vs_tg)} · vs LY {fmt_pct_html(runrate_vs_ly)}</div>
  </div>
  <div class="kpi-card bkk">
    <div class="kpi-label">BKK MTD</div>
    <div class="kpi-value">{fmt_m(bkk_mtd)}</div>
    <div class="kpi-sub">LY {fmt_m(bkk_mtd_ly)} · Diff <span class="{'pos' if bkk_diff>=0 else 'neg'}">{'+' if bkk_diff>=0 else ''}{fmt_m(bkk_diff)}</span> · vs TG {fmt_pct_html(bkk_vs_tg)} · vs LY {fmt_pct_html(bkk_vs_ly)}</div>
  </div>
  <div class="kpi-card upc">
    <div class="kpi-label">UPC MTD</div>
    <div class="kpi-value">{fmt_m(upc_mtd)}</div>
    <div class="kpi-sub">LY {fmt_m(upc_mtd_ly)} · Diff <span class="{'pos' if upc_diff>=0 else 'neg'}">{'+' if upc_diff>=0 else ''}{fmt_m(upc_diff)}</span> · vs TG {fmt_pct_html(upc_vs_tg)} · vs LY {fmt_pct_html(upc_vs_ly)}</div>
  </div>
  <div class="kpi-card mono">
    <div class="kpi-label">MONO MTD</div>
    <div class="kpi-value">{fmt_m(mono_mtd)}</div>
    <div class="kpi-sub">LY {fmt_m(mono_mtd_ly)} · Diff <span class="{'pos' if mono_diff>=0 else 'neg'}">{'+' if mono_diff>=0 else ''}{fmt_m(mono_diff)}</span> · vs TG(LY) {fmt_pct_html(mono_vs_tg)} · vs LY {fmt_pct_html(mono_vs_tg)}</div>
  </div>
</div>

<!-- PROGRESS BAR -->
<div class="progress-section">
  <div class="progress-label">SSP MTD vs Monthly Target &nbsp;<strong>{fmt_m(ssp_mtd)}</strong> / <strong>{fmt_m(ssp_monthly_target)}</strong></div>
  <div class="progress-track">
    <div class="progress-fill" style="width:{min(progress_pct, 100):.1f}%"></div>
    <div class="mtd-tg-marker" style="left:{min(mtd_tg_pct, 100):.1f}%" title="MTD Target: {fmt_m(mtd_target)}"></div>
  </div>
  <div class="progress-pct">{progress_pct:.0f}% &nbsp;<span style="color:#94a3b8;font-size:11px">MTD TG {fmt_m(mtd_target)} ({mtd_tg_pct:.0f}%)</span></div>
</div>

<!-- TABS -->
<div class="tabs">
  <button class="tab-btn active" onclick="showTab('summary',this)">Summary</button>
  <button class="tab-btn" onclick="showTab('ssp',this)">Supersports Stores</button>
  <button class="tab-btn" onclick="showTab('mono',this)">Mono-Brand Stores</button>
</div>

<!-- TAB: SUMMARY -->
<div id="tab-summary" class="tab-content active">

  <div class="card" style="margin-bottom:20px">
    <div class="section-title">Daily Sales Tracking (SSP — M฿)</div>
    <div class="table-wrap" style="max-height:520px;overflow-y:auto">
      <table class="daily-table">
        <thead><tr>
          <th>{now.year - 1}</th><th class="num">Net Sales</th>
          <th>{now.year}</th><th class="num">Daily Target</th>
          <th class="num">Actual Sales</th>
          <th class="num">Diff vs Target</th>
          <th class="num">%Chg vs LY</th>
          <th>Remark</th>
        </tr></thead>
        <tbody>{DAILY_TABLE_HTML}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- TAB: SSP STORES -->
<div id="tab-ssp" class="tab-content">
  <div class="filter-bar">{SSP_FILTER_BTNS}</div>
  <div class="table-wrap">
    <table id="ssp-table">
      <thead><tr>
        <th>Code</th><th>Store</th><th>DM</th><th>Region</th>
        <th class="num">Daily TY</th><th class="num">Daily LY</th><th class="num">Daily %</th>
        <th class="num">MTD TY</th><th class="num">MTD LY</th><th class="num">MTD TG</th><th class="num">%Ach</th><th class="num">MTD %LY</th>
      </tr></thead>
      <tbody id="ssp-body">{SSP_TABLE_HTML}</tbody>
    </table>
  </div>
</div>

<!-- TAB: MONO -->
<div id="tab-mono" class="tab-content">
  <div class="filter-bar">{MONO_FILTER_BTNS}</div>
  <div class="table-wrap">
    <table id="mono-table">
      <thead><tr>
        <th>Code</th><th>Store</th><th>Brand</th>
        <th class="num">Daily TY</th><th class="num">Daily LY</th><th class="num">Daily %</th>
        <th class="num">MTD TY</th><th class="num">MTD LY</th><th class="num">MTD %</th>
      </tr></thead>
      <tbody id="mono-body">{MONO_TABLE_HTML}</tbody>
    </table>
  </div>
</div>

<!-- SCRIPTS -->
<script>
function showTab(id,btn){{
  document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+id).classList.add('active');
  btn.classList.add('active');
}}
function filterSSP(region,btn){{
  document.querySelectorAll('#ssp-body tr').forEach(r=>{{
    if(region==='ALL'){{r.style.display='';return;}}
    const dr=r.dataset.region;
    r.style.display=(!dr||dr===region||dr==='ALL')?'':'none';
  }});
  btn.parentElement.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
}}
function filterMono(brand,btn){{
  document.querySelectorAll('#mono-body tr').forEach(r=>{{
    if(brand==='ALL'){{r.style.display='';return;}}
    const db=r.dataset.brand;
    r.style.display=(!db||db===brand)?'':'none';
  }});
  btn.parentElement.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
}}
</script>
</body>
</html>"""

# ═══════════════════════════════════════════════════════════════════════════════
# WRITE OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
out_name = f"DSR_Dashboard_{now.strftime('%b%Y')}.html"
out_path = Path(__file__).parent / out_name
out_path.write_text(html, encoding='utf-8')
print(f"Dashboard written → {out_path}")
print(f"  Data as of: {data_day_label}")
print(f"  SSP Daily: {fmt_m(ssp_daily)} | MTD: {fmt_m(ssp_mtd)} | Runrate: {fmt_m(runrate)}")
print(f"  BKK MTD: {fmt_m(bkk_mtd)} | UPC MTD: {fmt_m(upc_mtd)} | MONO MTD: {fmt_m(mono_mtd)}")

# ── Cleanup temp files ──
import tempfile, glob as _glob, os as _os
tmp_dir = tempfile.gettempdir()
cleaned = 0
script_dir = str(Path(__file__).parent)
for pat in [_os.path.join(tmp_dir, '.repaired_*'), _os.path.join(script_dir, '.repaired_*'), _os.path.join(script_dir, '**/daily_targets_cache.json')]:
    for f in _glob.glob(pat, recursive=True):
        try:
            _os.remove(f)
            cleaned += 1
        except OSError:
            pass
if cleaned:
    print(f"  Cleaned up {cleaned} temp file(s)")